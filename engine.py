import os
import json
import openpyxl
from datetime import datetime, date, timedelta
from typing import Dict, List, Any, Optional

EXCEL_PATH = r'C:\Users\siray\Downloads\Weekly Progress R2.xlsx'
CACHE_PATH = os.path.join(os.path.dirname(__file__), 'data_cache.json')
BACKUP_CACHE_PATH = os.path.join(os.path.dirname(__file__), 'data_cache_backup.json')
DEFAULT_WEBAPP_URL = 'https://script.google.com/macros/s/AKfycbwSbMxBfzkOWgXMA9OwZpu6-Y18Ap0mX1DFgXkZYvQ6P3NrKYpI4kKsxgz2LIEb6QmQ/exec'

def format_date(dt):
    if dt is None or dt == "-" or dt == "":
        return None
    if isinstance(dt, (datetime, date)):
        return dt.strftime('%Y-%m-%d')
    if isinstance(dt, str):
        try:
            return dt[:10]
        except:
            return None
    return None

def parse_date(d_str):
    if not d_str or d_str == "-":
        return None
    if isinstance(d_str, (datetime, date)):
        if isinstance(d_str, datetime):
            return d_str.date()
        return d_str
    if isinstance(d_str, str):
        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S'):
            try:
                return datetime.strptime(d_str.strip(), fmt).date()
            except ValueError:
                pass
    return None

class ProjectEngine:
    @staticmethod
    def is_cc_project(p: dict) -> bool:
        lot = str(p.get("lot", "")).strip().upper()
        name = str(p.get("name", "")).strip().upper()
        status = str(p.get("status", "")).strip().upper()
        return lot.startswith("CC") or " CC" in name or name.endswith("CC") or "ยกเลิก" in name or "CANCEL" in status

    def __init__(self, excel_path: str = EXCEL_PATH, cache_path: str = CACHE_PATH):
        self.excel_path = excel_path
        self.cache_path = cache_path
        self.backup_path = os.path.join(os.path.dirname(cache_path), 'data_cache_backup.json')
        self.weight_matrix = {}
        self.milestone_names = []
        self.milestone_categories = {}
        self.all_projects = []
        self.active_projects = []
        self.projects = []
        self.projects_dict = {}
        self.google_sheet_webapp_url = ''
        
        # Load from cache first
        if not self.load_from_cache():
            if os.path.exists(self.excel_path):
                self.load_data_from_excel()
                self.save_to_cache()
            else:
                print(f"[Engine Warning] Neither valid cache nor Excel file found.")

    def load_from_cache(self) -> bool:
        # Try primary cache
        for path in [self.cache_path, self.backup_path]:
            if os.path.exists(path):
                try:
                    with open(path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    
                    projects = data.get('projects', [])
                    if len(projects) > 0:
                        self.weight_matrix = {int(k): v for k, v in data.get('weight_matrix', {}).items()}
                        self.milestone_names = data.get('milestone_names', [])
                        self.milestone_categories = data.get('milestone_categories', {})
                        self.google_sheet_webapp_url = data.get('google_sheet_webapp_url') or DEFAULT_WEBAPP_URL
                        self.all_projects = projects
                        self.projects_dict = {p['id']: p for p in self.all_projects}
                        self.active_projects = [p for p in self.all_projects if not self.is_cc_project(p)]
                        self.projects = self.active_projects
                        print(f"[Fast Engine] Loaded {len(self.all_projects)} total projects ({len(self.active_projects)} active projects) successfully from {os.path.basename(path)}.")
                        return True
                except Exception as e:
                    print(f"[Engine] Error reading {path}: {e}")
        return False

    def save_to_cache(self):
        """
        Atomic cache saving to prevent any file corruption
        """
        try:
            self.active_projects = [p for p in self.all_projects if not self.is_cc_project(p)]
            self.projects = self.active_projects
            data = {
                'weight_matrix': self.weight_matrix,
                'milestone_names': self.milestone_names,
                'milestone_categories': self.milestone_categories,
                'projects': self.all_projects,
                'google_sheet_webapp_url': self.google_sheet_webapp_url
            }
            tmp_path = self.cache_path + '.tmp'
            with open(tmp_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False)
            
            # Atomic replace
            os.replace(tmp_path, self.cache_path)
            
            # Also keep a backup
            try:
                import shutil
                shutil.copyfile(self.cache_path, self.backup_path)
            except:
                pass
                
        except Exception as e:
            print(f"[Engine] Failed to save cache: {e}")

    def load_data_from_excel(self):
        if not os.path.exists(self.excel_path):
            print(f"File not found: {self.excel_path}")
            return

        print(f"[Engine] Reading Excel: {self.excel_path}...")
        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        self._load_weights(wb)
        self._load_projects(wb)
        self._calculate_all_scurves()

    def _load_weights(self, wb):
        if 'Weight Prj' not in wb.sheetnames:
            return
        ws = wb['Weight Prj']
        type_cols = {}
        for col in range(4, ws.max_column + 1):
            t_val = ws.cell(row=4, column=col).value
            if t_val is not None:
                try:
                    type_num = int(t_val)
                    type_cols[type_num] = col
                    if type_num not in self.weight_matrix:
                        self.weight_matrix[type_num] = {}
                except:
                    pass

        current_category = "งานขออนุญาตราชการ (Permission)"
        for r in range(6, ws.max_row + 1):
            c_name = ws.cell(row=r, column=2).value
            plan_act = ws.cell(row=r, column=3).value
            
            if c_name:
                c_name_str = str(c_name).strip()
                if "Engineering Design" in c_name_str:
                    current_category = "งานออกแบบวิศวกรรม (Engineering Design)"
                    continue
                elif "Construction Work" in c_name_str:
                    current_category = "งานก่อสร้างและติดตั้ง (Construction Work)"
                    continue
                elif "ราชการ" in c_name_str:
                    current_category = "งานขออนุญาตราชการ (Permission)"
                    continue
                
                if plan_act and str(plan_act).strip().upper() == "PLAN":
                    m_name = c_name_str
                    if m_name not in self.milestone_names:
                        self.milestone_names.append(m_name)
                        self.milestone_categories[m_name] = current_category
                    
                    for t_num, col_idx in type_cols.items():
                        w_val = ws.cell(row=r, column=col_idx).value
                        try:
                            w_float = float(w_val) if w_val is not None else 0.0
                        except:
                            w_float = 0.0
                        self.weight_matrix[t_num][m_name] = w_float

    def _load_projects(self, wb):
        if 'Plan' not in wb.sheetnames:
            return
        
        ws_plan = wb['Plan']
        ws_prog = wb['data Progress'] if 'data Progress' in wb.sheetnames else None
        
        plan_milestones_cols = []
        c = 8
        while c <= ws_plan.max_column:
            m_title = ws_plan.cell(row=3, column=c).value
            if not m_title:
                m_title = ws_plan.cell(row=2, column=c).value
            
            if m_title:
                m_title = str(m_title).strip()
                plan_milestones_cols.append({
                    "name": m_title,
                    "start_col": c,
                    "finish_col": c + 1,
                    "weight_col": c + 2
                })
                if m_title.lower() == "punch list":
                    break
                c += 3
            else:
                c += 1

        prog_data_by_prj = {}
        if ws_prog:
            for r in range(6, ws_prog.max_row + 1):
                p_name = ws_prog.cell(row=r, column=4).value
                if p_name:
                    p_name_str = str(p_name).strip()
                    prog_data_by_prj[p_name_str] = {}
                    for m_info in plan_milestones_cols:
                        m_name = m_info["name"]
                        s_c = m_info["start_col"]
                        f_c = m_info["finish_col"]
                        p_c = m_info["weight_col"]
                        
                        act_start = ws_prog.cell(row=r, column=s_c).value if s_c <= ws_prog.max_column else None
                        act_finish = ws_prog.cell(row=r, column=f_c).value if f_c <= ws_prog.max_column else None
                        act_pct = ws_prog.cell(row=r, column=p_c).value if p_c <= ws_prog.max_column else None
                        
                        try:
                            act_pct_val = float(act_pct) if act_pct is not None and str(act_pct).strip() not in ("", "-") else 0.0
                            if act_pct_val > 1.0:
                                act_pct_val = act_pct_val / 100.0
                        except:
                            act_pct_val = 0.0

                        prog_data_by_prj[p_name_str][m_name] = {
                            "actual_start": format_date(act_start),
                            "actual_finish": format_date(act_finish),
                            "actual_pct": act_pct_val
                        }

        self.projects = []
        self.projects_dict = {}
        self.google_sheet_webapp_url = ''
        
        for r in range(6, ws_plan.max_row + 1):
            p_name = ws_plan.cell(row=r, column=3).value
            if not p_name:
                continue
            p_name_str = str(p_name).strip()
            if not p_name_str or p_name_str == "None":
                continue
            
            bu = ws_plan.cell(row=r, column=1).value or "ทั่วไป"
            order_no = ws_plan.cell(row=r, column=2).value
            lot = ws_plan.cell(row=r, column=4).value or "Lot 1"
            capacity = ws_plan.cell(row=r, column=5).value or 0.0
            installation = ws_plan.cell(row=r, column=6).value or "Solar Rooftop"
            type_code = ws_plan.cell(row=r, column=7).value or 1
            
            try:
                capacity = float(capacity)
            except:
                capacity = 0.0
                
            try:
                type_code = int(type_code)
            except:
                type_code = 1

            project_id = f"prj_{r-5:03d}"
            milestones = []
            total_planned_weight = 0.0
            total_actual_progress = 0.0
            
            min_plan_start = None
            max_plan_finish = None
            min_act_start = None
            max_act_finish = None
            
            for m_info in plan_milestones_cols:
                m_name = m_info["name"]
                p_start = ws_plan.cell(row=r, column=m_info["start_col"]).value
                p_finish = ws_plan.cell(row=r, column=m_info["finish_col"]).value
                p_w = ws_plan.cell(row=r, column=m_info["weight_col"]).value
                
                try:
                    p_weight = float(p_w) if p_w is not None and str(p_w).strip() not in ("", "-") else 0.0
                except:
                    p_weight = self.weight_matrix.get(type_code, {}).get(m_name, 0.0)
                
                if p_weight == 0.0 and m_name in self.weight_matrix.get(type_code, {}):
                    p_weight = self.weight_matrix[type_code][m_name]
                    
                total_planned_weight += p_weight
                
                p_start_str = format_date(p_start)
                p_finish_str = format_date(p_finish)
                p_s_date = parse_date(p_start_str)
                p_f_date = parse_date(p_finish_str)
                
                if p_s_date:
                    if min_plan_start is None or p_s_date < min_plan_start:
                        min_plan_start = p_s_date
                if p_f_date:
                    if max_plan_finish is None or p_f_date > max_plan_finish:
                        max_plan_finish = p_f_date

                act_data = prog_data_by_prj.get(p_name_str, {}).get(m_name, {})
                act_start_str = act_data.get("actual_start")
                act_finish_str = act_data.get("actual_finish")
                act_pct = act_data.get("actual_pct", 0.0)
                
                if act_finish_str and act_pct == 0.0:
                    act_pct = 1.0
                elif act_start_str and not act_finish_str and act_pct == 0.0:
                    act_pct = 0.5

                a_s_date = parse_date(act_start_str)
                a_f_date = parse_date(act_finish_str)
                
                if a_s_date:
                    if min_act_start is None or a_s_date < min_act_start:
                        min_act_start = a_s_date
                if a_f_date:
                    if max_act_finish is None or a_f_date > max_act_finish:
                        max_act_finish = a_f_date

                milestone_actual_contrib = act_pct * p_weight
                total_actual_progress += milestone_actual_contrib
                category = self.milestone_categories.get(m_name, "งานทั่วไป")

                milestones.append({
                    "name": m_name,
                    "category": category,
                    "weight": round(p_weight, 4),
                    "planned_start": p_start_str,
                    "planned_finish": p_finish_str,
                    "actual_start": act_start_str,
                    "actual_finish": act_finish_str,
                    "actual_pct": round(act_pct, 4),
                    "actual_contribution": round(milestone_actual_contrib, 4),
                    "status": "COMPLETED" if act_pct >= 1.0 else ("IN_PROGRESS" if act_pct > 0 else "PENDING")
                })

            today = date.today()
            total_planned_progress_today = 0.0
            for m in milestones:
                p_s = parse_date(m["planned_start"])
                p_f = parse_date(m["planned_finish"])
                w = m["weight"]
                if p_s and p_f and w > 0:
                    if today >= p_f:
                        total_planned_progress_today += w
                    elif today <= p_s:
                        total_planned_progress_today += 0.0
                    else:
                        total_days = (p_f - p_s).days or 1
                        elapsed_days = (today - p_s).days
                        pct = min(1.0, max(0.0, elapsed_days / total_days))
                        total_planned_progress_today += pct * w

            actual_pct_total = min(100.0, total_actual_progress * 100)
            planned_pct_today = min(100.0, total_planned_progress_today * 100)
            diff = actual_pct_total - planned_pct_today
            
            if actual_pct_total >= 99.9:
                status = "COMPLETED"
                status_th = "เสร็จสมบูรณ์"
            elif diff >= 0:
                status = "ON_TRACK"
                status_th = "ตามแผนงาน"
            elif diff >= -10:
                status = "SLIGHT_DELAY"
                status_th = "ล่าช้าเล็กน้อย"
            else:
                status = "DELAYED"
                status_th = "ล่าช้ากว่าแผน"

            prj_obj = {
                "id": project_id,
                "business_unit": str(bu).strip(),
                "order_no": order_no,
                "name": p_name_str,
                "lot": str(lot).strip(),
                "capacity_kwp": round(capacity, 2),
                "installation_type": str(installation).strip(),
                "type_code": type_code,
                "planned_start": str(min_plan_start) if min_plan_start else None,
                "planned_finish": str(max_plan_finish) if max_plan_finish else None,
                "actual_start": str(min_act_start) if min_act_start else None,
                "actual_finish": str(max_act_finish) if max_act_finish else None,
                "actual_progress_pct": round(actual_pct_total, 2),
                "planned_progress_pct": round(planned_pct_today, 2),
                "variance_pct": round(diff, 2),
                "status": status,
                "status_th": status_th,
                "milestones": milestones
            }
            
            self.projects.append(prj_obj)
            self.projects_dict[project_id] = prj_obj

    def _calculate_all_scurves(self):
        for prj in self.projects:
            prj["s_curve"] = self.generate_project_scurve(prj)

    def generate_project_scurve(self, prj: Dict[str, Any]) -> Dict[str, Any]:
        milestones = prj.get("milestones", [])
        all_dates = []
        for m in milestones:
            for d_field in ["planned_start", "planned_finish", "actual_start", "actual_finish"]:
                dt = parse_date(m.get(d_field))
                if dt:
                    all_dates.append(dt)
        
        if not all_dates:
            return {"weeks": [], "labels": [], "planned_cum": [], "actual_cum": [], "planned_weekly": [], "actual_weekly": []}

        min_d = min(all_dates)
        max_d = max(all_dates)
        start_monday = min_d - timedelta(days=min_d.weekday())
        end_monday = max_d + timedelta(days=(7 - max_d.weekday()) % 7)
        if (end_monday - start_monday).days < 28:
            end_monday = start_monday + timedelta(days=35)

        weeks = []
        labels = []
        curr = start_monday
        w_idx = 1
        while curr <= end_monday:
            weeks.append(curr)
            labels.append(f"W{w_idx} ({curr.strftime('%d/%m/%y')})")
            curr += timedelta(days=7)
            w_idx += 1

        num_weeks = len(weeks)
        weekly_planned = [0.0] * num_weeks
        weekly_actual = [0.0] * num_weeks
        
        for m in milestones:
            w = m["weight"]
            if w <= 0:
                continue
            ps = parse_date(m.get("planned_start"))
            pf = parse_date(m.get("planned_finish"))
            if not ps or not pf:
                continue
            
            covered_indices = []
            for i, w_monday in enumerate(weeks):
                w_sunday = w_monday + timedelta(days=6)
                if not (pf < w_monday or ps > w_sunday):
                    covered_indices.append(i)
            
            if covered_indices:
                w_inc = w / len(covered_indices)
                for i in covered_indices:
                    weekly_planned[i] += w_inc

            act_pct = m.get("actual_pct", 0.0)
            if act_pct > 0:
                act_w = act_pct * w
                as_d = parse_date(m.get("actual_start")) or ps
                af_d = parse_date(m.get("actual_finish")) or date.today()
                
                act_covered_indices = []
                for i, w_monday in enumerate(weeks):
                    w_sunday = w_monday + timedelta(days=6)
                    if not (af_d < w_monday or as_d > w_sunday):
                        act_covered_indices.append(i)
                
                if act_covered_indices:
                    act_inc = act_w / len(act_covered_indices)
                    for i in act_covered_indices:
                        weekly_actual[i] += act_inc

        planned_cum = []
        actual_cum = []
        cum_p = 0.0
        cum_a = 0.0
        today = date.today()
        
        for i, w_monday in enumerate(weeks):
            cum_p += weekly_planned[i] * 100.0
            planned_cum.append(round(min(100.0, cum_p), 2))
            
            if w_monday <= today + timedelta(days=7):
                cum_a += weekly_actual[i] * 100.0
                actual_cum.append(round(min(100.0, cum_a), 2))
            else:
                actual_cum.append(None)

        return {
            "weeks": [w.strftime('%Y-%m-%d') for w in weeks],
            "labels": labels,
            "planned_cum": planned_cum,
            "actual_cum": actual_cum,
            "planned_weekly": [round(x * 100, 2) for x in weekly_planned],
            "actual_weekly": [round(x * 100, 2) for x in weekly_actual]
        }

    def get_phase_summary(self) -> List[Dict[str, Any]]:
        phases = {}
        for prj in self.projects:
            lot = prj.get("lot", "Other")
            if lot not in phases:
                phases[lot] = {
                    "lot": lot,
                    "project_count": 0,
                    "total_capacity_kwp": 0.0,
                    "actual_progress_weighted": 0.0,
                    "planned_progress_weighted": 0.0,
                    "projects": [],
                    "completed_count": 0,
                    "delayed_count": 0,
                    "on_track_count": 0
                }
            
            p = phases[lot]
            cap = prj.get("capacity_kwp", 0.0)
            p["project_count"] += 1
            p["total_capacity_kwp"] += cap
            p["actual_progress_weighted"] += prj.get("actual_progress_pct", 0.0) * cap
            p["planned_progress_weighted"] += prj.get("planned_progress_pct", 0.0) * cap
            p["projects"].append({
                "id": prj["id"],
                "name": prj["name"],
                "business_unit": prj["business_unit"],
                "capacity_kwp": cap,
                "installation_type": prj["installation_type"],
                "actual_progress_pct": prj["actual_progress_pct"],
                "planned_progress_pct": prj["planned_progress_pct"],
                "variance_pct": prj["variance_pct"],
                "status": prj["status"],
                "status_th": prj["status_th"]
            })
            
            if prj["status"] == "COMPLETED":
                p["completed_count"] += 1
            elif prj["status"] == "DELAYED":
                p["delayed_count"] += 1
            else:
                p["on_track_count"] += 1

        result = []
        for lot, data in phases.items():
            cap = data["total_capacity_kwp"]
            if cap > 0:
                data["avg_actual_progress"] = round(data["actual_progress_weighted"] / cap, 2)
                data["avg_planned_progress"] = round(data["planned_progress_weighted"] / cap, 2)
            else:
                data["avg_actual_progress"] = 0.0
                data["avg_planned_progress"] = 0.0
            data["total_capacity_kwp"] = round(data["total_capacity_kwp"], 2)
            result.append(data)
            
        return sorted(result, key=lambda x: x["lot"])

    def recalculate_project_metrics(self, prj: dict):
        """
        Recalculates progress %, status, variance %, and actual_start / actual_finish dates
        Rule: Project actual_finish is ONLY set if ALL milestones with weight > 0 are 100% completed.
        In that case, actual_finish = latest finish date of milestones with weight > 0.
        Otherwise, actual_finish = None (shows as '-').
        """
        milestones = prj.get("milestones", [])
        total_act = sum(m["actual_contribution"] for m in milestones)
        prj["actual_progress_pct"] = round(min(100.0, total_act * 100), 2)
        prj["variance_pct"] = round(prj["actual_progress_pct"] - prj["planned_progress_pct"], 2)
        
        if prj["actual_progress_pct"] >= 99.9:
            prj["status"] = "COMPLETED"
            prj["status_th"] = "เสร็จสมบูรณ์"
        elif prj["variance_pct"] >= 0:
            prj["status"] = "ON_TRACK"
            prj["status_th"] = "ตามแผนงาน"
        elif prj["variance_pct"] >= -10:
            prj["status"] = "SLIGHT_DELAY"
            prj["status_th"] = "ล่าช้าเล็กน้อย"
        else:
            prj["status"] = "DELAYED"
            prj["status_th"] = "ล่าช้ากว่าแผน"
            
        # Calculate actual_start: minimum start date of any started milestone
        start_dates = []
        for m in milestones:
            d = parse_date(m.get("actual_start"))
            if d:
                start_dates.append(d)
        prj["actual_start"] = min(start_dates).strftime('%Y-%m-%d') if start_dates else None
        
        # Calculate actual_finish: ONLY if all active milestones (weight > 0) are completed 100%
        active_milestones = [m for m in milestones if m.get("weight", 0) > 0]
        if active_milestones:
            all_active_completed = all(m.get("actual_pct", 0) >= 0.999 for m in active_milestones)
        else:
            all_active_completed = (prj["actual_progress_pct"] >= 99.9)
            
        if all_active_completed:
            finish_dates = []
            for m in active_milestones:
                d = parse_date(m.get("actual_finish"))
                if d:
                    finish_dates.append(d)
            prj["actual_finish"] = max(finish_dates).strftime('%Y-%m-%d') if finish_dates else date.today().strftime('%Y-%m-%d')
        else:
            prj["actual_finish"] = None
            
        prj["s_curve"] = self.generate_project_scurve(prj)

    def update_milestone(self, project_id: str, milestone_name: str = "", actual_pct: float = 0.0, 
                         actual_start: Optional[str] = None, actual_finish: Optional[str] = None,
                         milestone_index: Optional[int] = None) -> bool:
        if project_id not in self.projects_dict:
            return False
            
        prj = self.projects_dict[project_id]
        milestones = prj.get("milestones", [])
        target_m = None
        
        # 1. Match by milestone_index if provided
        if milestone_index is not None:
            try:
                idx = int(milestone_index)
                if 0 <= idx < len(milestones):
                    target_m = milestones[idx]
            except:
                pass
                
        # 2. Match by exact name
        if not target_m and milestone_name:
            clean_name = milestone_name.strip().lower()
            for m in milestones:
                if m["name"].strip().lower() == clean_name:
                    target_m = m
                    break
                    
        # 3. Match by normalized / whitespace-stripped name
        if not target_m and milestone_name:
            clean_nospace = "".join(milestone_name.lower().split())
            for m in milestones:
                m_clean = "".join(m["name"].lower().split())
                if m_clean == clean_nospace or m_clean in clean_nospace or clean_nospace in m_clean:
                    target_m = m
                    break
                    
        if not target_m:
            return False
            
        target_m["actual_pct"] = max(0.0, min(1.0, actual_pct))
        if actual_start:
            target_m["actual_start"] = actual_start
        if target_m["actual_pct"] >= 1.0:
            target_m["actual_finish"] = actual_finish or target_m.get("actual_finish") or date.today().strftime('%Y-%m-%d')
        elif actual_finish:
            target_m["actual_finish"] = actual_finish
        else:
            target_m["actual_finish"] = None
            
        target_m["status"] = "COMPLETED" if target_m["actual_pct"] >= 1.0 else ("IN_PROGRESS" if target_m["actual_pct"] > 0 else "PENDING")
        target_m["actual_contribution"] = round(target_m["actual_pct"] * target_m["weight"], 4)
        
        self.recalculate_project_metrics(prj)
        self.save_to_cache()
        return True

    def batch_sync_from_sheet_data(self, sheet_rows: list, m_names: list) -> int:
        updated_projects = 0
        for row in sheet_rows:
            if len(row) < 3:
                continue
            p_name = row[2].strip()
            if not p_name:
                continue
            
            p_order = row[1].strip() if len(row) > 1 else ''
            target_prj = None
            for p in self.all_projects:
                if p['name'].strip().lower() == p_name.lower() or (p_order and str(p.get('order_no')) == p_order):
                    target_prj = p
                    break
                    
            if not target_prj:
                continue
                
            m_idx = 0
            for c in range(7, len(row), 3):
                if m_idx >= len(m_names):
                    break
                m_name = m_names[m_idx]
                a_start = row[c].strip() if c < len(row) else ''
                a_finish = row[c+1].strip() if c+1 < len(row) else ''
                pct_raw = row[c+2].strip().replace('%', '') if c+2 < len(row) else '0'
                try:
                    val = float(pct_raw)
                    if val > 1.0:
                        val = val / 100.0
                except:
                    val = 0.0
                    
                for m in target_prj.get("milestones", []):
                    if m["name"].strip().lower() == m_name.strip().lower():
                        m["actual_pct"] = max(0.0, min(1.0, val))
                        if a_start:
                            m["actual_start"] = a_start
                        if val >= 1.0:
                            m["actual_finish"] = a_finish if a_finish else (m.get("actual_finish") or date.today().strftime('%Y-%m-%d'))
                        else:
                            m["actual_finish"] = None
                        m["status"] = "COMPLETED" if m["actual_pct"] >= 1.0 else ("IN_PROGRESS" if m["actual_pct"] > 0 else "PENDING")
                        m["actual_contribution"] = round(m["actual_pct"] * m["weight"], 4)
                        break
                        
                m_idx += 1
                
            self.recalculate_project_metrics(target_prj)
            updated_projects += 1
            
        self.save_to_cache()
        return updated_projects
